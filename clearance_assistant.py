
import clearance_assistant.utils as utils
import sys
from openpyxl import load_workbook
import openpyxl.styles as sty
import clearance_assistant.web_operator as web_operator

def usage():
    print('''
py clearance_assistannt.py <working_dir>

working_dir:仓管助手生成的结果文件夹。
''')


if len(sys.argv) == 1:
    print("\nError: 必须有1个参数用于指定文件夹名称")
    usage()
    exit(-1)

dir = sys.argv[1]
file = dir + "//结果//结果.xlsx"

wb = load_workbook(file)
ws = wb['半价清仓（可移仓）']

def process_meizhe():
    print("浏览器启动命令： chrome.exe --remote-debugging-port=9222 --user-data-dir=remote-profile")
    input("确保美折已在浏览器中登录。按任意键继续...")


    # 插入原价列和清仓价列
    orig_price_cn = utils.ws_get_column_cn(ws, "原价")
    if orig_price_cn == None:
        ws.cell(row=1, column=ws.max_column + 1).value = "原价"
        orig_price_cn = ws.max_column

    clearance_price_cn = utils.ws_get_column_cn(ws, "清仓价")
    if clearance_price_cn == None:
        ws.cell(row=1, column=ws.max_column + 1).value = "清仓价"
        clearance_price_cn = ws.max_column

    wo = web_operator.WebOperator()
    wo.meizhe_start_operation()

    for i in range(2, ws.max_row + 1):
        code = ws.cell(row=i, column=1).value

        # 仅处理清仓价不为空的行
        clearance_price = ws.cell(row=i, column=clearance_price_cn).value
        if clearance_price != None:
            continue

        print("正要设置：%d行，%s，按任意键继续：" % (i, code))
        input()

        ret = wo.meizhe_set_clearance_price_for_one_good(code)
        if ret == None:
            print("未找到商品")
            (orig_price, clearance_price) = ("未找到", "未找到")
        else:
            print("原价：%.2f，清仓价：%d" % (ret))
            (orig_price, clearance_price) = ret
        ws.cell(row=i, column=orig_price_cn).value = orig_price
        ws.cell(row=i, column=clearance_price_cn).value = clearance_price

        while (True):
            try:
                wb.save(file)
                break
            except IOError:
                input("文件无法保存，关掉其他应用后重试")


def process_cjdz():
    print("浏览器启动命令： chrome.exe --remote-debugging-port=9222 --user-data-dir=remote-profile")
    input("确保超级店长已在浏览器中登录。按任意键继续...")

    # excel表中插入“类目设置”列
    catagory_set_cn = utils.ws_get_column_cn(ws, "类目设置")
    if catagory_set_cn == None:
        ws.cell(row=1, column=ws.max_column + 1).value = "类目设置"
        catagory_set_cn = ws.max_column

    wo = web_operator.WebOperator()
    wo.cjdz_start_operation()

    for i in range(2, ws.max_row + 1):
        code = ws.cell(row=i, column=1).value

        # 仅处类目设置不为空的行
        catagory_set = ws.cell(row=i, column=catagory_set_cn).value
        if catagory_set != None:
            continue

        print("正要设置：%d行，%s，按任意键继续：" % (i, code))
        input()

        ret = wo.cjdz_check_one_good(code)

        v = "未找到"
        if ret == False:
            print("未找到商品")
        else:
            print("已勾选")
            v = 1

        ws.cell(row=i, column=catagory_set_cn).value = v

        while (True):
            try:
                wb.save(file)
                break
            except IOError:
                input("文件无法保存，关掉其他应用后重试")

def help():
    print(
"""命令帮助：
 h: 显示该帮助
 mz: 在美折中设置价格
 cjdz: 在超级店长中设置类目
 q: 退出
 """)


while True:
    cmd = input("输入命令(h：帮助)：")
    if cmd == "h":
        help()
    elif cmd == "mz":
        process_meizhe()

    elif cmd == "cjdz":
        process_cjdz()


    elif cmd == "q":
        break
    else:
        help()

wb.close()
exit(0)

